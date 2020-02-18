import datetime
import os
import re
import openpyxl

import pandas as pd
from flask import jsonify, request
from sqlalchemy import or_

from app import db
from app.api import bp
from app.models import WT, WTMaintain, User, Gzp, PowerCut
from app.api.users import get_user_id, get_user
from app.tool.tool import realRound
from app.api.dailyform import EXCEL_PATH


@bp.route('/createpc', methods=['POST'])
def create_pc():
    data = request.get_json() or {}
    # print(data)
    if 'id' in data:
        pc = PowerCut.query.filter_by(id=data['id']).first()
    else:
        pc = PowerCut()
    pc.stop_time = datetime.datetime.strptime(data['stop_time'], '%Y-%m-%d %H:%M')
    pc.start_time = datetime.datetime.strptime(data['start_time'], '%Y-%m-%d %H:%M')
    pc.lost_power1 = realRound(float(data['lost_power1']), 4)
    pc.lost_power2 = realRound(float(data['lost_power2']), 4)
    pc.time = realRound((pc.stop_time - pc.start_time).seconds / 3600, 2)
    db.session.add(pc)
    db.session.commit()
    return jsonify('ok')


@bp.route('/getpcs', methods=['GET'])
def get_pcs():
    pcs = PowerCut.query.filter(or_(PowerCut.start_time >= datetime.datetime.now().date(),
                                    PowerCut.stop_time >= datetime.datetime.now().date())).all()
    res = []
    for pc in pcs:
        res.append({
            'id': pc.id,
            'start_time': pc.start_time.strftime('%Y-%m-%d %H:%M'),
            'stop_time': pc.stop_time.strftime('%Y-%m-%d %H:%M'),
            'time': str(pc.time),
            'lost_power1': str(pc.lost_power1),
            'lost_power2': str(pc.lost_power2),
            'lost_power': str(realRound(pc.lost_power1 + pc.lost_power2, 4))

        })
    print(res)
    return jsonify(res)


@bp.route('/pctocdf', methods=['POST'])
def pc2cdf():
    """
    限电记录写入日报表
    """
    data = request.get_json() or {}
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    this_month = False
    flag = 0
    worksheet = workbook['电网故障、检修、限电统计']
    stop_time = datetime.datetime.strptime(data['stop_time'], '%Y-%m-%d %H:%M')
    start_time = datetime.datetime.strptime(data['start_time'], '%Y-%m-%d %H:%M')
    time = realRound((stop_time - start_time).seconds / 3600, 2)
    for row_num in range(1, worksheet.max_row):
        month = start_time.month
        if row_num == flag:
            continue
        if re.findall(r'石桥一期(\d)月电网故障/检修、电网限电统计', str(worksheet.cell(row_num, 1).value)):
            if str(month) == re.findall(r'石桥一期(\d)月电网故障/检修、电网限电统计', worksheet.cell(row_num, 1).value)[0]:
                #  定位到当月标题
                this_month = True
                flag = row_num + 1
        if this_month and worksheet.cell(row_num, 1).value in [None, '']:
            worksheet.cell(row_num, 1, '一二号集电线')
            worksheet.cell(row_num, 2, '电网限电')
            worksheet.cell(row_num, 3, 'A1-A6、A8-A11、A13-A20')
            worksheet.cell(row_num, 4, 'AGC曲线跟踪')
            worksheet.cell(row_num, 5, start_time)
            worksheet.cell(row_num, 6, stop_time)
            worksheet.cell(row_num, 7, time)
            worksheet.cell(row_num, 8, data['lost_power1'])
            worksheet.cell(row_num, 10, '三四号集电线')
            worksheet.cell(row_num, 11, '电网限电')
            worksheet.cell(row_num, 12, 'A21-A28、A30-A38、A40')
            worksheet.cell(row_num, 13, 'AGC曲线跟踪')
            worksheet.cell(row_num, 14, start_time)
            worksheet.cell(row_num, 15, stop_time)
            worksheet.cell(row_num, 16, time)
            worksheet.cell(row_num, 17, data['lost_power2'])
            break
    workbook.save(EXCEL_PATH)
    response = jsonify()
    response.status_code = 200
    # response.headers['Location'] = url_for('api.', id=task.id)
    return response


@bp.route('/changecdfpc', methods=['POST'])
def change_cdf_pc():
    """
    修改日报表中限电记录
    """
    data = request.get_json() or {}
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    response = jsonify()
    response.status_code = 202  # 未找到
    worksheet = workbook['电网故障、检修、限电统计']
    stop_time = datetime.datetime.strptime(data['new']['stop_time'], '%Y-%m-%d %H:%M')
    start_time = datetime.datetime.strptime(data['new']['start_time'], '%Y-%m-%d %H:%M')
    time = realRound((stop_time - start_time).seconds / 3600, 2)
    for row_num in range(1, worksheet.max_row):
        if type(worksheet.cell(row_num, 5).value) == datetime.datetime:  # 判断为时间类型才进入循环
            if worksheet.cell(row_num, 6).value.strftime('%Y-%m-%d %H:%M') == data['old']['stop_time'] \
                    and worksheet.cell(row_num, 5).value.strftime('%Y-%m-%d %H:%M') == \
                    data['old']['start_time']:
                worksheet.cell(row_num, 1, '一二号集电线')
                worksheet.cell(row_num, 2, '电网限电')
                worksheet.cell(row_num, 3, 'A1-A6、A8-A11、A13-A20')
                worksheet.cell(row_num, 4, 'AGC曲线跟踪')
                worksheet.cell(row_num, 5, start_time)
                worksheet.cell(row_num, 6, stop_time)
                worksheet.cell(row_num, 7, time)
                worksheet.cell(row_num, 8, data['new']['lost_power1'])
                worksheet.cell(row_num, 10, '三四号集电线')
                worksheet.cell(row_num, 11, '电网限电')
                worksheet.cell(row_num, 12, 'A21-A28、A30-A38、A40')
                worksheet.cell(row_num, 13, 'AGC曲线跟踪')
                worksheet.cell(row_num, 14, start_time)
                worksheet.cell(row_num, 15, stop_time)
                worksheet.cell(row_num, 16, time)
                worksheet.cell(row_num, 17, data['new']['lost_power2'])
                response.status_code = 200
                break
    workbook.save(EXCEL_PATH)
    return response


@bp.route('/delpcdb', methods=['POST'])
def del_pc_db():
    data = request.get_json() or {}
    pc = PowerCut.query.filter_by(id=data['id']).first()
    db.session.delete(pc)
    db.session.commit()
    return jsonify()

@bp.route('/delpccdf', methods=['POST'])
def del_pc_cdf():
    """
    删除日报表中限电记录
    """
    data = request.get_json() or {}
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    response = jsonify()
    response.status_code = 202  # 未找到
    worksheet = workbook['电网故障、检修、限电统计']
    for row_num in range(1, worksheet.max_row):
        if type(worksheet.cell(row_num, 5).value) == datetime.datetime:  # 判断为时间类型才进入循环
            if worksheet.cell(row_num, 6).value.strftime('%Y-%m-%d %H:%M') == data['stop_time'] \
                    and worksheet.cell(row_num, 5).value.strftime('%Y-%m-%d %H:%M') == \
                    data['start_time']:
                worksheet.cell(row_num, 1, '')
                worksheet.cell(row_num, 2, '')
                worksheet.cell(row_num, 3, '')
                worksheet.cell(row_num, 4, '')
                worksheet.cell(row_num, 5, '')
                worksheet.cell(row_num, 6, '')
                worksheet.cell(row_num, 7, '')
                worksheet.cell(row_num, 8, '')
                worksheet.cell(row_num, 10, '')
                worksheet.cell(row_num, 11, '')
                worksheet.cell(row_num, 12, '')
                worksheet.cell(row_num, 13, '')
                worksheet.cell(row_num, 14, '')
                worksheet.cell(row_num, 15, '')
                worksheet.cell(row_num, 16, '')
                worksheet.cell(row_num, 17, '')
                response.status_code = 200
                break
    workbook.save(EXCEL_PATH)
    return response
