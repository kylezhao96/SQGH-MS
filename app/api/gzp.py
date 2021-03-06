import datetime
import os
import re
from copy import copy

import numpy as np
import openpyxl

import pandas as pd
from flask import jsonify, request, make_response, send_file
from sqlalchemy.sql import and_, or_

from app import db
from app.api import bp
from app.models import WT, WTMaintain, User, Gzp
from app.api.users import get_user_id, get_user
from app.tool.tool import realRound, DESK_PATH
from app.api.dailyform import EXCEL_PATH


@bp.route('/createwtm', methods=['POST'])
def create_wtm():
    """
    使用此函数创建维护单
    """
    data = request.get_json() or {}
    print(data)
    # wtm = WTMaintain()
    gzp = Gzp()
    wtm = WTMaintain()
    # 暂时设定一张维护单对应一张工作票
    if not User.query.filter_by(name=data['manager']).first():  # 工作负责人不存在
        manager = User()
        manager.name = data['manager']
        db.session.add(manager)
        db.session.commit()
        gzp.manage_person_id = manager.id
    else:
        gzp.manage_person_id = User.query.filter_by(
            name=data['manager']).first().id
    wt_regex = re.compile(r'A(\d){,2}')
    gzp.wt_id = WT.query.filter_by(
        id=wt_regex.search(data['wt'][1]).group(1)).first().id
    gzp.task = data['task']
    wtm.type = data['type']
    # wtm.allow_time = datetime.datetime.fromtimestamp(data['allow_time']/1000)
    members = []
    for item in data['members']:
        if not User.query.filter_by(name=item).first():  # 工作班成员不存在
            member = User()
            member.name = item
            db.session.add(member)
            db.session.commit()
            members.append(members)
        else:
            members.append(User.query.filter_by(name=item).first())
    gzp.members = members
    db.session.add(wtm)
    db.session.commit()
    gzp.wtm_id = wtm.id
    db.session.add(gzp)
    db.session.commit()

    return jsonify("ok")


@bp.route('/getgzps', methods=['GET'])
def get_gzps():
    """
    获取所有今日工作票单及未终结工作票
    """
    today_gzps = Gzp.query.filter(or_(Gzp.pstart_time > datetime.date.today(), Gzp.is_end == False))
    # today_gzps = Gzp.query.order_by(-Gzp.pstart_time).limit(10).all()
    data = []
    for item in today_gzps:
        members = ''
        for member in item.members:
            members = members + member.name + '，'
        members = members[:-1]
        wts = ''
        wtms = list()
        index = 0
        for wt in item.wts:
            wtms.append({})
            if WTMaintain.query.filter_by(gzp_id=item.gzp_id, wt_id=wt.id).first():
                wtm = WTMaintain.query.filter_by(
                    gzp_id=item.gzp_id, wt_id=wt.id).first()
                wtms[index] = {
                    'wt_id': wt.id,
                    'stop_time': datetime.datetime.strftime(wtm.stop_time, '%Y-%m-%d %H:%M'),
                    'start_time': '' if not wtm.start_time else datetime.datetime.strftime(wtm.start_time,
                                                                                           '%Y-%m-%d %H:%M'),
                    'lost_power': wtm.lost_power,
                    'time': wtm.time
                }
            else:
                wtms[index] = {
                    'wt_id': wt.id,
                    'stop_time': '',
                    'start_time': '',
                    'lost_power': '',
                    'time': ''
                }
            wts = wts + 'A' + str(wt.id) + '，'
            index = index + 1
        wts = wts[:-1]
        x = {
            'id': item.gzp_id,
            'wt_id': wts,
            'manager': User.query.filter_by(id=item.manage_person_id).first().name,
            'task': item.task,
            'members': members,
            'pstart_time': datetime.datetime.strftime(item.pstart_time, '%Y-%m-%d %H:%M'),
            'pstop_time': datetime.datetime.strftime(item.pstop_time, '%Y-%m-%d %H:%M'),
            'wtms': wtms,
            'is_end': item.is_end,
        }
        if item.error_code:
            x['error_code'] = item.error_code
        data.append(x)
        index = index + 1
    print(data)
    return jsonify(data)


@bp.route('/getwttasks', methods=['GET'])
def get_wt_tasks():
    """
    获取所有创建过的任务
    """
    tasks = Gzp.query.with_entities(Gzp.task).distinct().all()
    re_tasks = []
    for item in tasks:
        x = {}
        re_tasks.append(
            {
                "value": item[0]
            }
        )
    print(re_tasks)
    return jsonify(re_tasks)


@bp.route('/delgzpdb', methods=['POST'])
def del_gzp_db():
    """
    删除工作票
    """
    data = request.get_json() or {}
    print(data)
    gzp = Gzp.query.filter_by(gzp_id=data['id']).first()
    db.session.delete(gzp)
    db.session.commit()
    response = jsonify()
    response.status_code = 200
    return response


@bp.route('/delgzpcdf', methods=['POST'])
def del_gzp_cdf():
    """
    删除日报表中维护记录
    """
    data = request.get_json() or {}
    wtms = data['wtms']
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    response = jsonify()
    response.status_code = 202  # 202代表未找到
    if 'error_code' in data:  # 故障
        wtm_type = '故障'
    else:
        wtm_type = '维护'
    worksheet = workbook['风机' + wtm_type + '统计']
    for wtm in wtms:
        if wtm['wt_id'] <= 20:  # 一号线
            #  在此写入
            start_col = 1
        elif wtm_type == '故障':  # 二号线 故障
            start_col = 11
        else:
            start_col = 9
        for col_num in range(1, worksheet.max_row):
            if type(worksheet.cell(col_num, start_col + 4).value) == datetime.datetime:  # 判断为时间类型才进入循环
                if wtm_type == '故障':
                    if worksheet.cell(col_num, start_col + 4).value.strftime('%Y-%m-%d %H:%M') == wtm['stop_time'] \
                            and worksheet.cell(col_num, start_col + 5).value.strftime('%Y-%m-%d %H:%M') == wtm[
                        'start_time']:
                        worksheet.cell(col_num, start_col, '')
                        worksheet.cell(col_num, start_col + 1, '')
                        worksheet.cell(col_num, start_col + 2, '')
                        worksheet.cell(col_num, start_col + 3, '')
                        worksheet.cell(col_num, start_col + 4, '')
                        worksheet.cell(col_num, start_col + 5, '')
                        worksheet.cell(col_num, start_col + 6, '')
                        worksheet.cell(col_num, start_col + 7, '')
                        worksheet.cell(col_num, start_col + 8, '')
                        response.status_code = 200  # 代表找到
                        break
                else:
                    if worksheet.cell(col_num, start_col + 3).value.strftime('%Y-%m-%d %H:%M') == wtm['stop_time'] \
                            and worksheet.cell(col_num, start_col + 4).value.strftime('%Y-%m-%d %H:%M') == wtm[
                        'start_time']:
                        worksheet.cell(col_num, start_col, '')
                        worksheet.cell(col_num, start_col + 1, '')
                        worksheet.cell(col_num, start_col + 2, '')
                        worksheet.cell(col_num, start_col + 3, '')
                        worksheet.cell(col_num, start_col + 4, '')
                        worksheet.cell(col_num, start_col + 5, '')
                        worksheet.cell(col_num, start_col + 6, '')
                        response.status_code = 200  # 代表找到
                        break
    if response.status_code == 200:
        try:
            workbook.save(EXCEL_PATH)
        except IOError:
            response.status_code = 501
    return response


@bp.route('/postgzp', methods=['POST'])
def post_gzp():
    data = request.files
    data.get('file').save('temp/gzp.xls')  # 将前端发来的文件暂存
    data_gzp = pd.read_excel('temp/gzp.xls')  # 读取
    if not Gzp.query.filter_by(gzp_id=data_gzp.loc[1].values[13]).first():
        gzp = Gzp()
    else:
        gzp = Gzp.query.filter_by(gzp_id=data_gzp.loc[1].values[13]).first()
    # 以下开始对各项数据进行读取
    gzp.firm = data_gzp.loc[1].values[1]
    gzp.gzp_id = data_gzp.loc[1].values[13]
    gzp.manage_person = get_user(data_gzp.loc[3].values[4])
    members = re.split("\W+", data_gzp.loc[6].values[0])
    members_temp = []
    for member in members:
        members_temp.append(get_user(member))
    gzp.members = members_temp
    if not pd.isnull(data_gzp.loc[9].values[5]):
        gzp.error_code = re.match(
            r'^(SC\d+_\d+_\d+)(\w+)?$', data_gzp.loc[9].values[5]).group(1)
        if re.match(r'^(SC\d+_\d+_\d+)(\w+)?$', data_gzp.loc[9].values[5]).group(2):
            gzp.error_content = re.match(
                r'^(SC\d+_\d+_\d+)(\w+)?$', data_gzp.loc[9].values[5]).group(2)
        else:
            gzp.error_content = re.match(
                r'(处理)?(\w+)', data_gzp.loc[11].values[10]).group(2)

    gzp_wts_id = list(
        map(lambda x: re.match(r'^(A)(\d+)$', x).group(2), re.findall(re.compile(r'A\d+'), data_gzp.loc[11].values[0])))
    gzp.wts = list(map(lambda x: WT.query.filter_by(
        id=int(x)).first(), gzp_wts_id))  # wt放在最后
    gzp.postion = data_gzp.loc[11].values[5]
    gzp.task = data_gzp.loc[11].values[10]
    index = 14
    gzp.pstart_time = datetime.datetime(data_gzp.loc[index + 1].values[2], data_gzp.loc[index + 1].values[4],
                                        data_gzp.loc[index + 1].values[6], data_gzp.loc[index + 1].values[8],
                                        data_gzp.loc[index + 1].values[10])
    gzp.pstop_time = datetime.datetime(data_gzp.loc[index + 2].values[2], data_gzp.loc[index + 2].values[4],
                                       data_gzp.loc[index + 2].values[6], data_gzp.loc[index + 2].values[8],
                                       data_gzp.loc[index + 2].values[10])
    # start_flag = 20
    # while True:  # flag 指向签发行
    #     if data_gzp.loc[start_flag].values[0] == '8、签发人签名':
    #         break
    #     start_flag = start_flag + 1
    # gzp.sign_person = get_user(data_gzp.loc[start_flag].values[2])
    # sign_time_year = data_gzp.loc[start_flag].values[7]
    # sign_time_month = data_gzp.loc[start_flag].values[9]
    # sign_time_day = data_gzp.loc[start_flag].values[11]
    # sign_time_hour = data_gzp.loc[start_flag].values[13]
    # sign_time_minutes = data_gzp.loc[start_flag].values[index+1]
    # gzp.sign_time = datetime.datetime(sign_time_year, sign_time_month, sign_time_day, sign_time_hour, sign_time_minutes)
    db.session.add(gzp)
    db.session.commit()

    os.remove('temp/gzp.xls')  # 删除暂存文件
    return jsonify('ok')


@bp.route('/wtmstodb', methods=['POST'])
def wtms2db():
    """
    将风机维护数据写入数据库
    """
    data = request.get_json() or {}
    gzp = Gzp.query.filter_by(gzp_id=data['id']).first()
    response = jsonify()
    for wtm in data['wtms']:
        if 'stop_time' in wtm.keys():
            if not WTMaintain.query.filter(WTMaintain.gzp_id == data['id'], WTMaintain.wt_id == wtm['wt_id']).first():
                wtm_db = WTMaintain()
            else:
                wtm_db = WTMaintain.query.filter(
                    WTMaintain.gzp_id == data['id'], WTMaintain.wt_id == wtm['wt_id']).first()
            wtm_db.wt_id = wtm['wt_id']
            gzp.wtms.append(wtm_db)
            gzp.is_end = True
            if gzp.error_code:  # 故障
                wtm_db.error_code = gzp.error_code
            else:
                wtm_db.task = gzp.task
            stop_time = datetime.datetime.strptime(
                wtm['stop_time'], '%Y-%m-%d %H:%M')
            wtm_db.stop_time = datetime.datetime(stop_time.year, stop_time.month, stop_time.day, stop_time.hour,
                                                 stop_time.minute)
            if 'start_time' in wtm.keys():
                if wtm['start_time']:
                    start_time = datetime.datetime.strptime(
                        wtm['start_time'], '%Y-%m-%d %H:%M')
                    wtm_db.start_time = datetime.datetime(start_time.year, start_time.month, start_time.day,
                                                          start_time.hour,
                                                          start_time.minute)
                    wtm_db.time = realRound(
                        (wtm_db.start_time - wtm_db.stop_time).seconds / 3600, 2)
                    wtm_db.lost_power = float(wtm['lost_power'])
                else:
                    gzp.is_end = False
                    response.status_code = 201
            else:
                gzp.is_end = False
                response.status_code = 201
            db.session.add(wtm_db)
            db.session.commit()
    db.session.add(gzp)
    db.session.commit()
    return response


@bp.route('/wtmstocdf', methods=['POST'])
def wtms2cdf():
    data = request.get_json() or {}
    gzp_id = data['gzp_id']
    response = jsonify()
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    gzp = Gzp.query.filter_by(gzp_id=gzp_id).first()
    this_month = False
    flag = 0
    if gzp.error_code:  # 故障
        wtm_type = '故障'
    else:
        wtm_type = '维护'
    worksheet = workbook['风机' + wtm_type + '统计']
    for wtm in gzp.wtms:
        if wtm.wt_id <= 20:  # 一号线
            #  在此写入
            start_col = 1
        elif wtm_type == '故障':  # 二号线 故障
            start_col = 11
        else:
            start_col = 9
        for col_num in range(1, worksheet.max_row):
            month = wtm.stop_time.month
            if col_num == flag:
                continue
            if re.findall(r'石桥一期(\d)月风机' + wtm_type + '统计', str(worksheet.cell(col_num, 1).value)):
                if str(month) == re.findall(r'石桥一期(\d)月风机' + wtm_type + '统计', worksheet.cell(col_num, 1).value)[0]:
                    #  定位到当月标题
                    this_month = True
                    flag = col_num + 1
            if this_month and worksheet.cell(col_num, 1).value == '合计':
                worksheet.insert_rows(col_num, amount=1)  # 若没有空行了，插入一行
                for x in range(1, 20):  # 复制样式
                    worksheet.cell(col_num, x)._style = copy(
                        worksheet.cell(col_num - 1, x)._style)
                    worksheet.cell(col_num, x).font = copy(
                        worksheet.cell(col_num - 1, x).font)
                    worksheet.cell(col_num, x).border = copy(
                        worksheet.cell(col_num - 1, x).border)
                    worksheet.cell(col_num, x).fill = copy(
                        worksheet.cell(col_num - 1, x).fill)
                    worksheet.cell(col_num, x).number_format = copy(
                        worksheet.cell(col_num - 1, x).number_format)
                    worksheet.cell(col_num, x).protection = copy(
                        worksheet.cell(col_num - 1, x).protection)
                    worksheet.cell(col_num, x).alignment = copy(
                        worksheet.cell(col_num - 1, x).alignment)
            if this_month and worksheet.cell(col_num, start_col).value in [None, '']:
                if wtm_type == '故障':
                    worksheet.cell(col_num, start_col,
                                   'A' + str(wtm.wt_id) + ' ' + str(WT.query.filter_by(id=wtm.wt_id).first().dcode))
                    worksheet.cell(col_num, start_col + 1, gzp.error_code)
                    worksheet.cell(col_num, start_col + 2, gzp.task)
                    worksheet.cell(col_num, start_col + 4, wtm.stop_time)
                    worksheet.cell(col_num, start_col + 5, wtm.start_time)
                    worksheet.cell(col_num, start_col + 6, wtm.time)
                    worksheet.cell(col_num, start_col + 7, wtm.lost_power)
                    # worksheet.cell(col_num, start_col+2, gzp.error_code)
                else:
                    worksheet.cell(col_num, start_col, 'A' + str(wtm.wt_id) + ' ' + str(
                        WT.query.filter_by(id=wtm.wt_id).first().dcode))
                    worksheet.cell(col_num, start_col + 1, '其他')
                    worksheet.cell(col_num, start_col + 2, gzp.task)
                    worksheet.cell(col_num, start_col + 3, wtm.stop_time)
                    worksheet.cell(col_num, start_col + 4, wtm.start_time)
                    worksheet.cell(col_num, start_col + 5, wtm.time)
                    worksheet.cell(col_num, start_col + 6, wtm.lost_power)
                break
    try:
        workbook.save(EXCEL_PATH)
    except IOError:
        response.status_code = 501
    return response


@bp.route('/changecdf', methods=['POST'])
def change_cdf():
    data = request.get_json() or {}
    wtms = data['new']['wtms']
    wtms_pre = data['old']['wtms']
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    response = jsonify()
    response.status_code = 201
    if 'error_code' in data['old']:  # 故障
        wtm_type = '故障'
    else:
        wtm_type = '维护'
    worksheet = workbook['风机' + wtm_type + '统计']
    for index, wtm in enumerate(wtms):
        stop_time = datetime.datetime.strptime(
            wtm['stop_time'], '%Y-%m-%d %H:%M')
        start_time = datetime.datetime.strptime(
            wtm['start_time'], '%Y-%m-%d %H:%M')
        if wtm['wt_id'] <= 20:  # 一号线
            #  在此写入
            start_col = 1
        elif wtm_type == '故障':  # 二号线 故障
            start_col = 11
        else:
            start_col = 9
        for col_num in range(1, worksheet.max_row):
            if type(worksheet.cell(col_num, start_col + 4).value) == datetime.datetime:  # 判断为时间类型才进入循环
                if wtm_type == '故障':
                    if worksheet.cell(col_num, start_col + 4).value.strftime('%Y-%m-%d %H:%M') == wtms_pre[index][
                        'stop_time'] \
                            and worksheet.cell(col_num, start_col + 5).value.strftime('%Y-%m-%d %H:%M') == \
                            wtms_pre[index][
                                'start_time']:
                        worksheet.cell(col_num, start_col,
                                       'A' + str(wtm['wt_id']) + ' ' + str(
                                           WT.query.filter_by(id=wtm['wt_id']).first().dcode))
                        worksheet.cell(col_num, start_col + 1,
                                       data['new']['error_code'])
                        worksheet.cell(col_num, start_col +
                                       2, data['new']['task'])
                        worksheet.cell(col_num, start_col + 4, stop_time)
                        worksheet.cell(col_num, start_col + 5, start_time)
                        worksheet.cell(
                            col_num, start_col + 6, realRound((start_time - stop_time).seconds / 3600, 2))
                        worksheet.cell(col_num, start_col +
                                       7, wtm['lost_power'])
                        # worksheet.cell(col_num, start_col+2, gzp.error_code)
                        response.status_code = 200
                        break
                else:
                    if worksheet.cell(col_num, start_col + 3).value.strftime('%Y-%m-%d %H:%M') == wtms_pre[index][
                        'stop_time'] \
                            and worksheet.cell(col_num, start_col + 4).value.strftime('%Y-%m-%d %H:%M') == \
                            wtms_pre[index][
                                'start_time']:
                        worksheet.cell(col_num, start_col, 'A' + str(wtm['wt_id']) + ' ' + str(
                            WT.query.filter_by(id=wtm['wt_id']).first().dcode))
                        worksheet.cell(col_num, start_col + 1, '其他')
                        worksheet.cell(col_num, start_col +
                                       2, data['new']['task'])
                        worksheet.cell(col_num, start_col + 3, stop_time)
                        worksheet.cell(col_num, start_col + 4, start_time)
                        worksheet.cell(
                            col_num, start_col + 5, realRound((start_time - stop_time).seconds / 3600, 2))
                        worksheet.cell(col_num, start_col +
                                       6, wtm['lost_power'])
                        response.status_code = 200
                        break
    if response.status_code == 200:
        try:
            workbook.save(EXCEL_PATH)
        except IOError:
            response.status_code = 501
    return response


@bp.route('/wtmsyn', methods=['GET'])
def wtm_syn():
    res = gzp_syn()  # 现将工作票同步
    whtj = pd.read_excel(EXCEL_PATH, sheet_name='风机维护统计',
                         usecols=range(20), header=None).fillna('')
    gztj = pd.read_excel(EXCEL_PATH, sheet_name='风机故障统计',
                         usecols=range(20), header=None).fillna('')
    # 维护
    for x in range(len(whtj)):
        if re.findall(r'^(A)(\d+)(\s*)(\d{5})$', whtj.loc[x].values[0]):
            stop_time = whtj.loc[x].values[3]
            zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                       seconds=stop_time.second)
            wt_id = int(re.match(r'^(A)(\d+)(\s*)(\d{5})$', whtj.loc[x].values[0]).group(2))

            gzp = Gzp.query.filter(
                and_(Gzp.pstart_time < stop_time, Gzp.pstart_time > zero_time,
                     Gzp.wts.any(id=wt_id))).first()  # 这里可能会生成bug
            wtm = WTMaintain()
            is_in = False
            for item in gzp.wtms:
                if item == WTMaintain.query.filter(WTMaintain.stop_time == whtj.loc[x].values[3],
                                                   WTMaintain.start_time == whtj.loc[x].values[4]).first():
                    wtm = item
                    is_in = True  # 标定数据库中已存在
            wtm.wt_id = wt_id
            wtm.type = whtj.loc[x].values[1]
            wtm.task = whtj.loc[x].values[2]
            wtm.stop_time = whtj.loc[x].values[3]
            wtm.start_time = whtj.loc[x].values[4]
            wtm.time = realRound(
                (wtm.start_time - wtm.stop_time).seconds / 3600, 2)
            wtm.lost_power = realRound(float(whtj.loc[x].values[6]), 4)
            if not is_in:
                gzp.wtms.append(wtm)
            if len(gzp.wtms.all()) == len(gzp.wts.all()):
                gzp.is_end = 1
            db.session.add(gzp)
            db.session.commit()
        if re.findall(r'^(A)(\d+)(\s*)(\d{5})$', whtj.loc[x].values[8]):
            stop_time = whtj.loc[x].values[11]
            zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                       seconds=stop_time.second)
            wt_id = int(re.match(r'^(A)(\d+)(\s*)(\d{5})$', whtj.loc[x].values[8]).group(2))
            gzp = Gzp.query.filter(
                and_(Gzp.pstart_time < stop_time, Gzp.pstart_time > zero_time,
                     Gzp.wts.any(id=wt_id))).first()  # 这里可能会生成bug
            wtm = WTMaintain()
            is_in = False
            for item in gzp.wtms:
                if item == WTMaintain.query.filter(WTMaintain.stop_time == whtj.loc[x].values[11],
                                                   WTMaintain.start_time == whtj.loc[x].values[12]).first():
                    wtm = item
                    is_in = True  # 标定数据库中已存在
            wtm.wt_id = wt_id
            wtm.type = whtj.loc[x].values[9]
            wtm.task = whtj.loc[x].values[10]
            wtm.stop_time = whtj.loc[x].values[11]
            wtm.start_time = whtj.loc[x].values[12]
            wtm.time = realRound(
                (wtm.start_time - wtm.stop_time).seconds / 3600, 2)
            wtm.lost_power = realRound(float(whtj.loc[x].values[13]), 4)
            if not is_in:
                gzp.wtms.append(wtm)
            if len(gzp.wtms.all()) == len(gzp.wts.all()):
                gzp.is_end = 1
            db.session.add(gzp)
            db.session.commit()
    # 故障
    for x in range(len(gztj)):
        if re.findall(r'^(A)(\d+)(\s*)(\d{5})$', gztj.loc[x].values[0]):
            stop_time = gztj.loc[x].values[4]
            start_time = gztj.loc[x].values[5]
            if stop_time.hour < 18:
                zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                           seconds=stop_time.second) + datetime.timedelta(days=1)
            else:
                zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                           seconds=stop_time.second) + datetime.timedelta(days=2)
            wt_id = int(re.match(r'^(A)(\d+)(\s*)(\d{5})$', gztj.loc[x].values[0]).group(2))
            try:
                gzp = Gzp.query.filter(
                    and_(Gzp.pstart_time > stop_time, Gzp.pstart_time < zero_time,
                         Gzp.wts.any(id=wt_id))).first()  # 这里可能会生成bug
                wtm = WTMaintain()
                is_in = False
                for item in gzp.wtms:
                    if item == WTMaintain.query.filter(WTMaintain.stop_time == stop_time,
                                                       WTMaintain.start_time == start_time).first():
                        wtm = item
                        is_in = True  # 标定数据库中已存在
                wtm.wt_id = wt_id
                wtm.error_code = gztj.loc[x].values[1]
                wtm.error_content = gztj.loc[x].values[2]
                wtm.type = gztj.loc[x].values[3]
                wtm.stop_time = gztj.loc[x].values[4]
                wtm.start_time = gztj.loc[x].values[5]
                wtm.time = realRound(
                    (wtm.start_time - wtm.stop_time).seconds / 3600, 2)
                wtm.lost_power = realRound(float(gztj.loc[x].values[7]), 4)
                wtm.error_approach = gztj.loc[x].values[8]
                wtm.task = gzp.task
                if not is_in:
                    gzp.wtms.append(wtm)
                if len(gzp.wtms.all()) == len(gzp.wts.all()):
                    gzp.is_end = 1
                db.session.add(gzp)
                db.session.commit()
            except(AttributeError):
                print('工作票不存在，风机号：A' + str(wt_id) + '，停机时间' + str(stop_time))
                res.append('风机号：A' + str(wt_id) + '，停机时间' + str(stop_time))
        if re.findall(r'^(A)(\d+)(\s*)(\d{5})$', gztj.loc[x].values[10]):
            stop_time = gztj.loc[x].values[14]
            start_time = gztj.loc[x].values[15]
            if stop_time.hour < 18:
                zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                           seconds=stop_time.second) + datetime.timedelta(days=1)
            else:
                zero_time = stop_time - datetime.timedelta(hours=stop_time.hour, minutes=stop_time.minute,
                                                           seconds=stop_time.second) + datetime.timedelta(days=2)
            wt_id = int(re.match(r'^(A)(\d+)(\s*)(\d{5})$', gztj.loc[x].values[10]).group(2))
            try:
                gzp = Gzp.query.filter(
                    and_(Gzp.pstart_time > stop_time, Gzp.pstart_time < zero_time,
                         Gzp.wts.any(id=wt_id))).first()  # 这里可能会生成bug
                wtm = WTMaintain()
                is_in = False
                for item in gzp.wtms:
                    if item == WTMaintain.query.filter(WTMaintain.stop_time == stop_time,
                                                       WTMaintain.start_time == start_time).first():
                        wtm = item
                        is_in = True  # 标定数据库中已存在
                wtm.wt_id = wt_id
                wtm.error_code = gztj.loc[x].values[11]
                wtm.error_content = gztj.loc[x].values[12]
                wtm.type = gztj.loc[x].values[13]
                wtm.stop_time = gztj.loc[x].values[14]
                wtm.start_time = gztj.loc[x].values[15]
                wtm.time = realRound(
                    (wtm.start_time - wtm.stop_time).seconds / 3600, 2)
                wtm.lost_power = realRound(float(gztj.loc[x].values[17]), 4)
                wtm.error_approach = gztj.loc[x].values[18]
                wtm.task = gzp.task
                if not is_in:
                    gzp.wtms.append(wtm)
                if len(gzp.wtms.all()) == len(gzp.wts.all()):
                    gzp.is_end = 1
                db.session.add(gzp)
                db.session.commit()
            except(AttributeError):
                print('工作票不存在，风机号：A' + str(wt_id) + '，停机时间' + str(stop_time))
                res.append('风机号：A' + str(wt_id) + '，停机时间' + str(stop_time))
    return jsonify(res)


# @bp.route('/gzpsyn', methods=['GET'])
def gzp_syn():
    # data = request.get_json() or {}
    res = []
    path = DESK_PATH + r'5OA系统风机工作票'
    for year_folder in os.listdir(path):
        if re.match('\d+年$', year_folder):
            for month_folder in os.listdir(path + '\\' + year_folder):
                if re.match('\d+月$', month_folder):
                    for gzp in os.listdir(path + '\\' + year_folder + '\\' + month_folder):
                        if re.match(r'^(风机检修工作票)\S+(\.xls)$', gzp):
                            data_gzp = pd.read_excel(
                                path + '\\' + year_folder + '\\' + month_folder + '\\' + gzp)  # 读取
                            if not Gzp.query.filter_by(gzp_id=data_gzp.loc[1].values[13]).first():
                                gzp = Gzp()
                            else:
                                gzp = Gzp.query.filter_by(
                                    gzp_id=data_gzp.loc[1].values[13]).first()
                            # 公司
                            gzp.firm = data_gzp.loc[1].values[1]
                            # 以下开始对各项数据进行读取
                            for index, row in data_gzp.fillna('').iterrows():
                                for col_num in range(0, 19):
                                    if row.values[col_num] != '' and isinstance(row.values[col_num], str):
                                        # 匹配编号
                                        if re.match(r'\S{4}-\S{2}-\S{2}-\d{9}', row.values[col_num]):
                                            gzp.gzp_id = re.match(r'\S{4}-\S{2}-\S{2}-\d{9}',
                                                                  row.values[col_num]).group()
                                        # 匹配工作负责人
                                        if row.values[col_num] == '1、工作负责人(监护人)':
                                            gzp.manage_person = get_user(row.values[col_num + 4])
                                        # 匹配工作班成员
                                        if row.values[col_num] == '2、工作班成员（不包括工作负责人）':
                                            members = re.split(
                                                "\W+", data_gzp.loc[index + 1].values[0].strip())
                                            members_temp = []
                                            for member in members:
                                                members_temp.append(get_user(member))
                                            gzp.members = members_temp
                                        # 匹配故障
                                        if row.values[col_num] == '3、工作任务':
                                            if isinstance(data_gzp.loc[index + 1].values[5], str):
                                                if re.match(r'(SC\d+_\d+_\d+)\\?(\w+)?',
                                                            data_gzp.loc[index + 1].values[5]):
                                                    gzp.error_code = re.match(
                                                        r'(SC\d+_\d+_\d+)\\?(\w+)?',
                                                        data_gzp.loc[index + 1].values[5]).group(1)
                                                    if re.match(r'(SC\d+_\d+_\d+)\\?(\w+)?',
                                                                data_gzp.loc[index + 1].values[5]).group(2):
                                                        gzp.error_content = re.match(
                                                            r'(SC\d+_\d+_\d+)\\?(\w+)?',
                                                            data_gzp.loc[index + 1].values[5]).group(2)
                                                    else:
                                                        gzp.error_content = re.match(
                                                            r'(处理)?(\w+)', data_gzp.loc[index + 3].values[10]).group(2)
                                            # 匹配风机
                                            gzp_wts_id = list(
                                                map(lambda x: re.match(r'^(A)(\d+)$', x).group(2),
                                                    re.findall(re.compile(r'A\d+'), data_gzp.loc[index + 3].values[0])))
                                            gzp.wts = list(map(lambda x: WT.query.filter_by(
                                                id=int(x)).first(), gzp_wts_id))  # wt放在最后
                                            gzp.postion = data_gzp.loc[index + 3].values[5]
                                            gzp.task = data_gzp.loc[index + 3].values[10]
                                        # 匹配时间
                                        if row.values[col_num] == '5、计划工作时间':
                                            try:
                                                gzp.pstart_time = datetime.datetime(data_gzp.loc[index + 1].values[2],
                                                                                    data_gzp.loc[index + 1].values[4],
                                                                                    data_gzp.loc[index + 1].values[6],
                                                                                    data_gzp.loc[index + 1].values[8],
                                                                                    data_gzp.loc[index + 1].values[10])
                                                now = datetime.datetime.now()
                                                zeroToday = now - datetime.timedelta(hours=now.hour, minutes=now.minute,
                                                                                     seconds=now.second,
                                                                                     microseconds=now.microsecond)

                                                if gzp.pstart_time < zeroToday:
                                                    gzp.is_end = 1  # 读取到非今日工作票记为已终结
                                                if data_gzp.loc[index + 2].values[8] != 24:
                                                    gzp.pstop_time = datetime.datetime(
                                                        data_gzp.loc[index + 2].values[2],
                                                        data_gzp.loc[index + 2].values[4],
                                                        data_gzp.loc[index + 2].values[6],
                                                        data_gzp.loc[index + 2].values[8],
                                                        data_gzp.loc[index + 2].values[10])
                                                else:
                                                    gzp.pstop_time = datetime.datetime(
                                                        data_gzp.loc[index + 2].values[2],
                                                        data_gzp.loc[index + 2].values[4],
                                                        data_gzp.loc[index + 2].values[6], 0,
                                                        data_gzp.loc[index + 2].values[10]) + datetime.timedelta(days=1)
                                            except ValueError:
                                                res.append(gzp.gzp_id)
                                                print(gzp.gzp_id)
                            # start_flag = 20
                            # while True:  # flag 指向签发行
                            #     if data_gzp.loc[start_flag].values[0] == '8、签发人签名':
                            #         break
                            #     start_flag = start_flag + 1
                            # gzp.sign_person = get_user(data_gzp.loc[start_flag].values[2])
                            # sign_time_year = data_gzp.loc[start_flag].values[7]
                            # sign_time_month = data_gzp.loc[start_flag].values[9]
                            # sign_time_day = data_gzp.loc[start_flag].values[11]
                            # sign_time_hour = data_gzp.loc[start_flag].values[13]
                            # sign_time_minutes = data_gzp.loc[start_flag].values[index+1]
                            # gzp.sign_time = datetime.datetime(sign_time_year, sign_time_month, sign_time_day,
                            #                                   sign_time_hour, sign_time_minutes)
                            db.session.add(gzp)
    db.session.commit()
    return res


@bp.route('/stat2excel', methods=['GET'])
def gzp_by_users():
    wb = openpyxl.Workbook()
    users = User.query.all()
    ft = openpyxl.styles.Font(bold=True)
    alignment = openpyxl.styles.Alignment(horizontal='center',
                                          # 水平'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
                                          vertical='center',  # 垂直'distributed', 'bottom', 'top', 'center', 'justify'
                                          text_rotation=0,  # 旋转角度0~180
                                          wrap_text=True,  # 文字换行
                                          shrink_to_fit=False,  # 自适应宽度，改变文字大小,上一项false
                                          indent=0)
    for user in users:
        # 工作班成员
        if len(user.gzps.all()):
            wb.create_sheet(title=user.name)
            ws = wb[user.name]
            ws.column_dimensions['A'].width = 25
            ws.cell(1, 1).value = '票号'
            ws.column_dimensions['B'].width = 20
            ws.cell(1, 2).value = '时间'
            ws.column_dimensions['C'].width = 20
            ws.cell(1, 3).value = '风机号'
            ws.column_dimensions['D'].width = 40
            ws.cell(1, 4).value = '工作内容'
            ws.column_dimensions['D'].width = 20
            ws.cell(1, 5).value = '工作负责人'
            col_num = 2
            for gzp in user.gzps.all() + Gzp.query.filter(Gzp.manage_person == user).all():
                ws.cell(col_num, 1).value = gzp.gzp_id
                ws.cell(col_num, 2).value = gzp.pstart_time.date()
                wts = ''
                for wt in gzp.wts:
                    wts = wts + 'A' + str(wt.id) + '、'
                wts = wts[:-1]
                ws.cell(col_num, 3).value = wts
                ws.cell(col_num, 4).value = gzp.task
                ws.cell(col_num, 5).value = gzp.manage_person.name
                for irow, row in enumerate(ws.rows, start=1):
                    for cell in row:
                        cell.alignment = alignment
                        if irow == 1:
                            cell.font = ft
                col_num = col_num + 1
    ws = wb['Sheet']
    wb.remove(ws)
    wb.save(DESK_PATH + r'工作票工作成员统计.xlsx')
    return jsonify({})


@bp.route('/gzp_by_days', methods=['POST'])
def gzp_by_days():
    data = request.get_json() or {}
    print(data)
    startTime = datetime.datetime.strptime(data['startTime'], '%Y-%m-%d')
    endTime = datetime.datetime.strptime(data['endTime'], '%Y-%m-%d')
    users_id = data['users']
    users= User.query.filter(User.id.in_(users_id)).all()
    gzps = Gzp.query.filter(Gzp.pstart_time >= startTime, Gzp.pstop_time <= endTime).all()
    res = []
    for gzp in gzps:
        if gzp.manage_person_id in users_id or (set(gzp.members) & set(users)):
            res.append({
                'gzp_id': gzp.gzp_id,
                'wt': wtsToString(gzp.wts),
                'date': gzp.pstart_time.date().strftime('%Y-%m-%d'),
                'task': gzp.task,
                'manage_person': gzp.manage_person.name,
                'member': memberToString(gzp.members.all())
            })
    return jsonify(res)


def wtsToString(wts):
    wtString = ''
    for wt in wts:
        wtString = wtString + 'A' + str(wt.id) + '，'
    return wtString[:-1]


def memberToString(members):
    memberString = ''
    for member in members:
        memberString = memberString + member.name + '，'
    return memberString[:-1]


@bp.route('/getGzpAnalysisExcel',methods=['POST'])
def getGzpAnalysisExcel():
    gzps= request.get_json() or {}
    wb = openpyxl.Workbook()
    ft = openpyxl.styles.Font(bold=True)
    alignment = openpyxl.styles.Alignment(horizontal='center',
                                          # 水平'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
                                          vertical='center',  # 垂直'distributed', 'bottom', 'top', 'center', 'justify'
                                          text_rotation=0,  # 旋转角度0~180
                                          wrap_text=True,  # 文字换行
                                          shrink_to_fit=False,  # 自适应宽度，改变文字大小,上一项false
                                          indent=0)
    ws = wb.active
    ws.column_dimensions['A'].width = 25
    ws.cell(1, 1).value = '票号'
    ws.column_dimensions['B'].width = 20
    ws.cell(1, 2).value = '时间'
    ws.column_dimensions['C'].width = 20
    ws.cell(1, 3).value = '风机号'
    ws.column_dimensions['D'].width = 40
    ws.cell(1, 4).value = '工作内容'
    ws.column_dimensions['E'].width = 20
    ws.cell(1, 5).value = '工作负责人'
    ws.column_dimensions['F'].width = 40
    ws.cell(1, 6).value = '工作班成员'
    col_num = 2
    for gzp in gzps:
        ws.cell(col_num, 1).value = gzp['gzp_id']
        ws.cell(col_num, 2).value = gzp['date']
        ws.cell(col_num, 3).value = gzp['wt']
        ws.cell(col_num, 4).value = gzp['task']
        ws.cell(col_num, 5).value = gzp['manage_person']
        ws.cell(col_num, 6).value = gzp['member']
        for irow, row in enumerate(ws.rows, start=1):
            for cell in row:
                cell.alignment = alignment
                if irow == 1:
                    cell.font = ft
        col_num = col_num + 1
    wb.save(r'temp\gzpAnalysis.xlsx')
    response = make_response(send_file('D:\\MyRepositories\\sqgh-ms\\temp\\gzpAnalysis.xlsx'))
    response.headers["Content-Disposition"] = "attachment; filename= gzpAnalysis.xlsx;"
    return response