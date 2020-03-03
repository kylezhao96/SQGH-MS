import datetime
import json
from decimal import Decimal

import openpyxl
# from flask_cors import cross_origin
import pandas as pd
import pythoncom
import xlwings as xl
from flask import jsonify, request
from sqlalchemy import or_

from app import db
from app.api import bp
from app.models import CalDailyForm, WTMaintain, PowerCut
from app.tool.tool import realRound, DecimalEncoder, OMS_PATH, get_stop_time, get_lost_power, TY_PATH, get_num,EXCEL_PATH



def save_excel(path):
    """
    对打开的excel文件进行保存
    """
    pythoncom.CoInitialize()
    app = xl.App(visible=False)
    book = app.books.open(path)
    book.save()
    app.kill()


@bp.route('/cdfsyn', methods=['GET'])
def import_cdf():
    """
    自动化对日报表进行读取同步
    """
    cdf = pd.read_excel(EXCEL_PATH, sheet_name='日报计算表', usecols=range(76), skiprows=range(3), header=None)
    ty = pd.read_excel(TY_PATH, sheet_name='风速统计', usecols=range(3), skiprows=range(2), header=None).fillna('')
    cdf.fillna(0)
    print(cdf)
    response = []
    this_year = datetime.date.today().year
    for x in range(366):
        # if CalDailyForm.query.filter_by(date=cdf.loc[x + 1].values[0] + datetime.timedelta(-1)).first():
        # continue  # 如果数据库存在本日数据，那么跳过
        # cdf_db = CalDailyForm.query.filter_by(date=cdf.loc[x + 1].values[0] + datetime.timedelta(-1)).first()
        data = {}
        if x == 0:
            data['date'] = datetime.datetime(this_year - 1, 12, 31, 0, 0)
            data['fka312'] = float(cdf.loc[x].values[1])
            data['bka312'] = float(cdf.loc[x].values[2])
            data['fka313'] = float(cdf.loc[x].values[3])
            data['bka313'] = float(cdf.loc[x].values[4])
            data['fka322'] = float(cdf.loc[x].values[5])
            data['bka322'] = float(cdf.loc[x].values[6])
            data['fka323'] = float(cdf.loc[x].values[7])
            data['bka323'] = float(cdf.loc[x].values[8])
            data['fka31b'] = float(cdf.loc[x].values[9])
            data['fka21b'] = float(cdf.loc[x].values[10])
            data['fka311'] = float(cdf.loc[x].values[11])
            data['bka311'] = float(cdf.loc[x].values[12])
            data['fkr311'] = float(cdf.loc[x].values[13])
            data['bkr311'] = float(cdf.loc[x].values[14])
            data['fka321'] = float(cdf.loc[x].values[15])
            data['bka321'] = float(cdf.loc[x].values[16])
            data['fkr321'] = float(cdf.loc[x].values[17])
            data['bkr321'] = float(cdf.loc[x].values[18])
            data['bka111'] = float(cdf.loc[x].values[19])
            data['fka111'] = float(cdf.loc[x].values[20])
        else:
            if pd.isnull(cdf.loc[x].values[1]):
                # if cdf.loc[x].values[0] >= datetime.now():
                break  # 如果读到的数据不是浮点数类型，那么停止
            data['date'] = cdf.loc[x].values[0]
            data['fka312'] = float(cdf.loc[x].values[1])
            data['bka312'] = float(cdf.loc[x].values[2])
            data['fka313'] = float(cdf.loc[x].values[3])
            data['bka313'] = float(cdf.loc[x].values[4])
            data['fka322'] = float(cdf.loc[x].values[5])
            data['bka322'] = float(cdf.loc[x].values[6])
            data['fka323'] = float(cdf.loc[x].values[7])
            data['bka323'] = float(cdf.loc[x].values[8])
            data['fka31b'] = float(cdf.loc[x].values[9])
            data['fka21b'] = float(cdf.loc[x].values[10])
            data['fka311'] = float(cdf.loc[x].values[11])
            data['bka311'] = float(cdf.loc[x].values[12])
            data['fkr311'] = float(cdf.loc[x].values[13])
            data['bkr311'] = float(cdf.loc[x].values[14])
            data['fka321'] = float(cdf.loc[x].values[15])
            data['bka321'] = float(cdf.loc[x].values[16])
            data['fkr321'] = float(cdf.loc[x].values[17])
            data['bkr321'] = float(cdf.loc[x].values[18])
            data['bka111'] = float(cdf.loc[x].values[19])
            data['fka111'] = float(cdf.loc[x].values[20])
            data['dgp1'] = cdf.loc[x].values[21]
            data['donp1'] = cdf.loc[x].values[22]
            data['doffp1'] = cdf.loc[x].values[23]
            data['dcp1'] = cdf.loc[x].values[24]
            data['dcl1'] = cdf.loc[x].values[25]
            data['dgp2'] = cdf.loc[x].values[26]
            data['donp2'] = cdf.loc[x].values[27]
            data['doffp2'] = cdf.loc[x].values[28]
            data['dcp2'] = cdf.loc[x].values[29]
            data['dcl2'] = cdf.loc[x].values[30]
            data['dgp'] = cdf.loc[x].values[31]
            data['donp'] = cdf.loc[x].values[32]
            data['doffp'] = cdf.loc[x].values[33]
            data['dcp'] = cdf.loc[x].values[34]
            data['dcl'] = cdf.loc[x].values[35]
            data['doffp31b'] = cdf.loc[x].values[36]
            data['doffp21b'] = cdf.loc[x].values[37]
            data['agp1'] = cdf.loc[x].values[38]
            data['aonp1'] = cdf.loc[x].values[39]
            data['aoffp1'] = cdf.loc[x].values[40]
            data['acp1'] = cdf.loc[x].values[41]
            data['acl1'] = cdf.loc[x].values[42]
            data['agp2'] = cdf.loc[x].values[43]
            data['aonp2'] = cdf.loc[x].values[44]
            data['aoffp2'] = cdf.loc[x].values[45]
            data['acp2'] = cdf.loc[x].values[46]
            data['acl2'] = cdf.loc[x].values[47]
            data['agp'] = cdf.loc[x].values[48]
            data['aonp'] = cdf.loc[x].values[49]
            data['aoffp'] = cdf.loc[x].values[50]
            data['acp'] = cdf.loc[x].values[51]
            data['acl'] = cdf.loc[x].values[52]
            data['mgp1'] = cdf.loc[x].values[53]
            data['monp1'] = cdf.loc[x].values[54]
            data['moffp1'] = cdf.loc[x].values[55]
            data['mcp1'] = cdf.loc[x].values[56]
            data['mcl1'] = cdf.loc[x].values[57]
            data['mgp2'] = cdf.loc[x].values[58]
            data['monp2'] = cdf.loc[x].values[59]
            data['moffp2'] = cdf.loc[x].values[60]
            data['mcp2'] = cdf.loc[x].values[61]
            data['mcl2'] = cdf.loc[x].values[62]
            data['mgp'] = cdf.loc[x].values[63]
            data['monp'] = cdf.loc[x].values[64]
            data['moffp'] = cdf.loc[x].values[65]
            data['mcp'] = cdf.loc[x].values[66]
            data['mcl'] = cdf.loc[x].values[67]
            data['offja311'] = cdf.loc[x].values[69]
            data['offjr311'] = cdf.loc[x].values[71]
            data['offja321'] = cdf.loc[x].values[73]
            data['offjr321'] = cdf.loc[x].values[75]
            data['davgs1'] = float(ty.loc[x - 1].values[1])
            data['davgs2'] = float(ty.loc[x - 1].values[2])
            data['davgs'] = float(realRound((data['davgs1'] + data['davgs2']) / 2, 2))
        response.append(data)
        if CalDailyForm.query.filter_by(date=cdf.loc[x + 1].values[0] + datetime.timedelta(-1)).first():
            # continue  # 如果数据库存在本日数据，那么跳过
            cdf_db = CalDailyForm.query.filter_by(date=cdf.loc[x + 1].values[0] + datetime.timedelta(-1)).first()
            cdf_db.from_dict(data)
            db.session.add(cdf_db)
            db.session.commit()
        else:
            cdf2 = CalDailyForm()
            cdf2.from_dict(data)
            db.session.add(cdf2)
            db.session.commit()
    return jsonify(response)


@bp.route('/addtodaycdf', methods=["POST"])
def add_cdf(day):
    """
    将日报表计算出的输入日期的表码值填写入数据库中
    """
    cdf = pd.read_excel(EXCEL_PATH, sheet_name='日报计算表', usecols=range(76), skiprows=range(3), header=None)
    data = {}
    x = (day - datetime.datetime(datetime.datetime.now().year, 1, 1)).days + 1
    if not CalDailyForm.query.filter_by(date=cdf.loc[x].values[0]).first():
        data['date'] = cdf.loc[x].values[0]
        data['fka312'] = float(cdf.loc[x].values[1])
        data['bka312'] = float(cdf.loc[x].values[2])
        data['fka313'] = float(cdf.loc[x].values[3])
        data['bka313'] = float(cdf.loc[x].values[4])
        data['fka322'] = float(cdf.loc[x].values[5])
        data['bka322'] = float(cdf.loc[x].values[6])
        data['fka323'] = float(cdf.loc[x].values[7])
        data['bka323'] = float(cdf.loc[x].values[8])
        data['fka31b'] = float(cdf.loc[x].values[9])
        data['fka21b'] = float(cdf.loc[x].values[10])
        data['fka311'] = float(cdf.loc[x].values[11])
        data['bka311'] = float(cdf.loc[x].values[12])
        data['fkr311'] = float(cdf.loc[x].values[13])
        data['bkr311'] = float(cdf.loc[x].values[14])
        data['fka321'] = float(cdf.loc[x].values[15])
        data['bka321'] = float(cdf.loc[x].values[16])
        data['fkr321'] = float(cdf.loc[x].values[17])
        data['bkr321'] = float(cdf.loc[x].values[18])
        data['bka111'] = float(cdf.loc[x].values[19])
        data['fka111'] = float(cdf.loc[x].values[20])
        data['dgp1'] = cdf.loc[x].values[21]
        data['donp1'] = cdf.loc[x].values[22]
        data['doffp1'] = cdf.loc[x].values[23]
        data['dcp1'] = cdf.loc[x].values[24]
        data['dcl1'] = cdf.loc[x].values[25]
        data['dgp2'] = cdf.loc[x].values[26]
        data['donp2'] = cdf.loc[x].values[27]
        data['doffp2'] = cdf.loc[x].values[28]
        data['dcp2'] = cdf.loc[x].values[29]
        data['dcl2'] = cdf.loc[x].values[30]
        data['dgp'] = cdf.loc[x].values[31]
        data['donp'] = cdf.loc[x].values[32]
        data['doffp'] = cdf.loc[x].values[33]
        data['dcp'] = cdf.loc[x].values[34]
        data['dcl'] = cdf.loc[x].values[35]
        data['doffp31b'] = cdf.loc[x].values[36]
        data['doffp21b'] = cdf.loc[x].values[37]
        data['agp1'] = cdf.loc[x].values[38]
        data['aonp1'] = cdf.loc[x].values[39]
        data['aoffp1'] = cdf.loc[x].values[40]
        data['acp1'] = cdf.loc[x].values[41]
        data['acl1'] = cdf.loc[x].values[42]
        data['agp2'] = cdf.loc[x].values[43]
        data['aonp2'] = cdf.loc[x].values[44]
        data['aoffp2'] = cdf.loc[x].values[45]
        data['acp2'] = cdf.loc[x].values[46]
        data['acl2'] = cdf.loc[x].values[47]
        data['agp'] = cdf.loc[x].values[48]
        data['aonp'] = cdf.loc[x].values[49]
        data['aoffp'] = cdf.loc[x].values[50]
        data['acp'] = cdf.loc[x].values[51]
        data['acl'] = cdf.loc[x].values[52]
        data['mgp1'] = cdf.loc[x].values[53]
        data['monp1'] = cdf.loc[x].values[54]
        data['moffp1'] = cdf.loc[x].values[55]
        data['mcp1'] = cdf.loc[x].values[56]
        data['mcl1'] = cdf.loc[x].values[57]
        data['mgp2'] = cdf.loc[x].values[58]
        data['monp2'] = cdf.loc[x].values[59]
        data['moffp2'] = cdf.loc[x].values[60]
        data['mcp2'] = cdf.loc[x].values[61]
        data['mcl2'] = cdf.loc[x].values[62]
        data['mgp'] = cdf.loc[x].values[63]
        data['monp'] = cdf.loc[x].values[64]
        data['moffp'] = cdf.loc[x].values[65]
        data['mcp'] = cdf.loc[x].values[66]
        data['mcl'] = cdf.loc[x].values[67]
        data['offja311'] = cdf.loc[x].values[69]
        data['offjr311'] = cdf.loc[x].values[71]
        data['offja321'] = cdf.loc[x].values[73]
        data['offjr321'] = cdf.loc[x].values[75]

        cdf2 = CalDailyForm()
        cdf2.from_dict(data)
        db.session.add(cdf2)
        db.session.commit()


@bp.route('/toexcel', methods=["POST"])
def fill_caldailyform():
    """
    将表码值录入日报表
    """
    data = request.get_json() or {}
    # data['data'] = datetime.datetime.combine(datetime.date.today(), datetime.time(0, 0, 0))
    # print(data)
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    worksheet = workbook['日报计算表']
    rowNum = (getdate() - datetime.datetime(datetime.datetime.now().year, 1, 1)).days + 5
    # rowNum = (datetime.datetime(2019, 10, 3) - datetime.datetime(datetime.datetime.now().year, 1, 1)).days+5
    worksheet.cell(rowNum, 2, data['fka312'])  # 312正向
    worksheet.cell(rowNum, 3, data['bka312'])  # 312反向
    worksheet.cell(rowNum, 4, data['fka313'])  # 313正向
    worksheet.cell(rowNum, 5, data['bka313'])  # 313反向
    worksheet.cell(rowNum, 6, data['fka322'])  # 322正向
    worksheet.cell(rowNum, 7, data['bka322'])  # 322反向
    worksheet.cell(rowNum, 8, data['fka323'])  # 323正向
    worksheet.cell(rowNum, 9, data['bka323'])  # 323正向
    worksheet.cell(rowNum, 10, data['fka31b'])  # 31b正向
    worksheet.cell(rowNum, 11, data['fka21b'])  # 21b正向
    worksheet.cell(rowNum, 12, 0)  # 311正向有
    worksheet.cell(rowNum, 13, data['bka311'])  # 311反向有
    worksheet.cell(rowNum, 14, 0)  # 311正向无
    worksheet.cell(rowNum, 15, data['bkr311'])  # 311反向无
    worksheet.cell(rowNum, 16, 0)  # 321正向有
    worksheet.cell(rowNum, 17, data['bka321'])  # 321反向有
    worksheet.cell(rowNum, 18, 0)  # 321正向无
    worksheet.cell(rowNum, 19, data['bkr321'])  # 321反向无
    worksheet.cell(rowNum, 20, data['bka111'])  # 111反向
    worksheet.cell(rowNum, 21, data['fka111'])  # 111正向
    workbook.save(EXCEL_PATH)
    save_excel(EXCEL_PATH)
    response = jsonify(data)
    response.status_code = 200
    return response


def getdate():
    """
    用于判断日报时间的函数
    """
    this_hour = datetime.datetime.now().hour
    if 0 <= this_hour <= 8:
        return datetime.datetime.combine(datetime.date.today(), datetime.time(0, 0, 0)) + datetime.timedelta(-1)
    else:
        return datetime.datetime.combine(datetime.date.today(), datetime.time(0, 0, 0))


def calcdf(data):
    """
    根据表码值计算电量
    """
    cdf_pre = CalDailyForm.query.filter_by(date=data['date'] + datetime.timedelta(-1)).first()  # 前一日表码值
    if data['date'].day == 1:  # 1日月归零
        cdf_pre.mgp1 = 0
        cdf_pre.mgp2 = 0
        cdf_pre.mgp = 0
        cdf_pre.monp1 = 0
        cdf_pre.monp2 = 0
        cdf_pre.monp = 0
        cdf_pre.moffp1 = 0
        cdf_pre.moffp2 = 0
        cdf_pre.moffp = 0
        cdf_pre.mcp1 = 0
        cdf_pre.mcp2 = 0
        cdf_pre.mcp = 0
        cdf_pre.mlp1 = 0
        cdf_pre.mlp2 = 0
        cdf_pre.mlp = 0
        if [data['date'].month == 1]:  # 1月1日年归零
            cdf_pre.agp1 = 0
            cdf_pre.agp2 = 0
            cdf_pre.agp = 0
            cdf_pre.aonp1 = 0
            cdf_pre.aonp2 = 0
            cdf_pre.aonp = 0
            cdf_pre.aoffp1 = 0
            cdf_pre.aoffp2 = 0
            cdf_pre.aoffp = 0
            cdf_pre.acp1 = 0
            cdf_pre.acp2 = 0
            cdf_pre.acp = 0
            cdf_pre.alp1 = 0
            cdf_pre.alp2 = 0
            cdf_pre.alp = 0
    res = {}
    # 以下为日数据
    res['dgp1'] = realRound(
        (float(data['fka312']) - cdf_pre.fka312 + float(data['fka313']) - cdf_pre.fka313) * 42000 * 1.01)
    res['dgp2'] = realRound(
        (float(data['fka322']) - cdf_pre.fka322 + float(data['fka323']) - cdf_pre.fka323) * 42000 * 1.01)
    res['dgp'] = res['dgp1'] + res['dgp2']
    res['donp1'] = realRound(
        0 if res['dgp1'] == 0 else Decimal(float(data['bka111']) - cdf_pre.bka111) * 176000 * res['dgp1'] / res['dgp'])
    res['donp2'] = realRound(
        0 if res['dgp1'] == 0 else Decimal(float(data['bka111']) - cdf_pre.bka111) * 176000 * res['dgp2'] / res['dgp'])
    res['donp'] = res['donp1'] + res['donp2']
    bka_diff_1 = float(data['bka312']) + float(data['bka313']) - cdf_pre.bka312 - cdf_pre.bka313
    bka_diff_2 = float(data['bka322']) + float(data['bka323']) - cdf_pre.bka322 - cdf_pre.bka323
    res['doffp1'] = realRound(
        0 if bka_diff_2 + bka_diff_1 == 0 else (float(data['fka111']) - cdf_pre.fka111) * 176000 * bka_diff_1 / (
                bka_diff_1 + bka_diff_2))
    res['doffp2'] = realRound(
        0 if bka_diff_2 + bka_diff_1 == 0 else (float(data['fka111']) - cdf_pre.fka111) * 176000 * bka_diff_2 / (
                bka_diff_1 + bka_diff_2))
    res['doffp'] = res['doffp1'] + res['doffp2']
    res['doffp31b'] = realRound((float(data['fka31b']) - cdf_pre.fka31b) * 7000)
    res['doffp21b'] = realRound((float(data['fka21b']) - cdf_pre.fka21b) * 120)
    res['dcp1'] = res['dgp1'] - res['donp1'] + res['doffp1'] + res['doffp21b']
    res['dcp2'] = res['dgp2'] - res['donp2'] + res['doffp2']
    res['dcp'] = res['dcp1'] + res['dcp2']
    res['dcl1'] = realRound(0 if res['dgp1'] == 0 else res['dcp1'] / (res['dgp1']), 4)  # 这里思考为保存两位小数，因为计算公式较为简单
    res['dcl2'] = realRound(0 if res['dgp2'] == 0 else res['dcp2'] / (res['dgp2']), 4)
    res['dcl'] = realRound(0 if res['dgp'] == 0 else res['dcp'] / (res['dgp']), 4)
    res['offja311'] = (float(data['bka311']) - cdf_pre.bka311) * 21000
    res['offjr311'] = (float(data['bkr311']) - cdf_pre.bkr311) * 21000
    res['offja321'] = (float(data['bka321']) - cdf_pre.bka321) * 21000
    res['offjr321'] = (float(data['bkr321']) - cdf_pre.bkr321) * 21000
    # 以下为月数据
    res['mgp1'] = res['dgp1'] + Decimal(cdf_pre.mgp1)
    res['mgp2'] = res['dgp2'] + Decimal(cdf_pre.mgp2)
    res['mgp'] = res['dgp'] + Decimal(cdf_pre.mgp)
    res['monp1'] = res['donp1'] + Decimal(cdf_pre.monp1)
    res['monp2'] = res['donp2'] + Decimal(cdf_pre.monp2)
    res['monp'] = res['donp'] + Decimal(cdf_pre.monp)
    res['moffp1'] = res['doffp1'] + Decimal(cdf_pre.moffp1)
    res['moffp2'] = res['doffp2'] + Decimal(cdf_pre.moffp2)
    res['moffp'] = res['doffp'] + Decimal(cdf_pre.moffp)
    res['mcp1'] = res['dcp1'] + Decimal(cdf_pre.mcp1)
    res['mcp2'] = res['dcp2'] + Decimal(cdf_pre.mcp2)
    res['mcp'] = res['dcp'] + Decimal(cdf_pre.mcp)
    res['mcl1'] = realRound(0 if res['mgp1'] == 0 else res['mcp1'] / (res['mgp1']), 4)  # 这里思考为保存两位小数，因为计算公式较为简单
    res['mcl2'] = realRound(0 if res['mgp2'] == 0 else res['mcp2'] / (res['mgp2']), 4)
    res['mcl'] = realRound(0 if res['mgp'] == 0 else res['mcp'] / (res['mgp']), 4)
    # 以下为年数据
    res['agp1'] = res['dgp1'] + Decimal(cdf_pre.agp1)
    res['agp2'] = res['dgp2'] + Decimal(cdf_pre.agp2)
    res['agp'] = res['dgp'] + Decimal(cdf_pre.agp)
    res['aonp1'] = res['donp1'] + Decimal(cdf_pre.aonp1)
    res['aonp2'] = res['donp2'] + Decimal(cdf_pre.aonp2)
    res['aonp'] = res['donp'] + Decimal(cdf_pre.aonp)
    res['aoffp1'] = res['doffp1'] + Decimal(cdf_pre.aoffp1)
    res['aoffp2'] = res['doffp2'] + Decimal(cdf_pre.aoffp2)
    res['aoffp'] = res['doffp'] + Decimal(cdf_pre.aoffp)
    res['acp1'] = res['dcp1'] + Decimal(cdf_pre.acp1)
    res['acp2'] = res['dcp2'] + Decimal(cdf_pre.acp2)
    res['acp'] = res['dcp'] + Decimal(cdf_pre.acp)
    res['acl1'] = realRound(0 if res['agp1'] == 0 else res['acp1'] / (res['agp1']), 4)  # 这里思考为保存两位小数，因为计算公式较为简单
    res['acl2'] = realRound(0 if res['agp2'] == 0 else res['acp2'] / (res['agp2']), 4)
    res['acl'] = realRound(0 if res['agp'] == 0 else res['acp'] / (res['agp']), 4)
    # 限电计算
    pcs = PowerCut.query.filter(or_(PowerCut.start_time >= datetime.datetime.now().date(),
                                    PowerCut.stop_time >= datetime.datetime.now().date())).all()
    res['dlp1'] = 0
    res['dlp2'] = 0
    res['dlp'] = 0
    for pc in pcs:
        res['dlp1'] = realRound(res['dlp1'] + Decimal(pc.lost_power1), 4)
        res['dlp2'] = realRound(res['dlp2'] + Decimal(pc.lost_power2), 4)
    res['dlp'] = res['dlp1'] + res['dlp2']
    res['mlp1'] = res['dlp1'] + Decimal(cdf_pre.mlp1)
    res['mlp2'] = res['dlp2'] + Decimal(cdf_pre.mlp2)
    res['mlp'] = res['dlp'] + Decimal(cdf_pre.mlp)
    res['alp1'] = res['dlp1'] + Decimal(cdf_pre.alp1)
    res['alp2'] = res['dlp2'] + Decimal(cdf_pre.alp2)
    res['alp'] = res['dlp'] + Decimal(cdf_pre.alp)
    return res


@bp.route('/filldb', methods=["POST"])
def fill_db():
    """
    将填写数据写入数据库
    """
    data = request.get_json() or {}
    print(data)
    data['date'] = getdate()
    data.update(calcdf(data))
    data['dmins'] = min(float(data['dmins1']), float(data['dmins2']))
    data['dmaxs'] = max(float(data['dmaxs1']), float(data['dmaxs2']))
    data['davgs'] = realRound((float(data['davgs1']) + float(data['davgs2'])) / 2, 2)
    if not CalDailyForm.query.filter_by(date=getdate()).first():
        cdf = CalDailyForm()
        cdf.from_dict(data)
        db.session.add(cdf)
        db.session.commit()
    else:
        cdf = CalDailyForm.query.filter_by(date=getdate()).first()
        cdf.from_dict(data)
        db.session.commit()
    # cdf = CalDailyForm.query.filter_by(date=getdate()).first()
    data['dlpl1'] = realRound(data['dlp1'] / data['dgp1'], 4)  # 由于限电率不保存入数据库，故写入数据库完计算
    data['dlpl2'] = realRound(data['dlp2'] / data['dgp2'], 4)
    data['dlpl'] = realRound(data['dlp'] / data['dgp'], 4)
    data['mlpl1'] = realRound(data['mlp1'] / data['mgp1'], 4)  # 由于限电率不保存入数据库，故写入数据库完计算
    data['mlpl2'] = realRound(data['mlp2'] / data['mgp2'], 4)
    data['mlpl'] = realRound(data['mlp'] / data['mgp'], 4)
    data['alpl1'] = realRound(data['alp1'] / data['agp1'], 4)  # 由于限电率不保存入数据库，故写入数据库完计算
    data['alpl2'] = realRound(data['alp2'] / data['agp2'], 4)
    data['alpl'] = realRound(data['alp'] / data['agp'], 4)
    res, flag = [], -1
    sets = [['日', 'd', '', 1], ['一期', 'd', '1', 11], ['二期', 'd', '2', 12], ['月', 'm', '', 2], ['一期', 'm', '1', 21],
            ['二期', 'm', '2', 22], ['年', 'a', '', 3], ['一期', 'a', '1', 31], ['二期', 'a', '2', 32]]
    for set in sets:
        elem = {
            'type': set[0],
            'gp': realRound((data[set[1] + 'gp' + set[2]]) / 10000, 4),
            'onp': realRound((data[set[1] + 'onp' + set[2]]) / 10000, 4),
            'offp': realRound((data[set[1] + 'offp' + set[2]]) / 10000, 4),
            'cp': realRound((data[set[1] + 'cp' + set[2]]) / 10000, 4),
            'cl': format(data[set[1] + 'cl' + set[2]] * 100, '0.2f') + '%',
            'id': set[3],
            'lp': data[set[1] + 'lp' + set[2]],
            'lpl': format(data[set[1] + 'lpl' + set[2]] * 100, '0.2f') + '%',
        }
        if set[1] == 'd':
            elem['mins'] = data['dmins' + set[2]]
            elem['maxs'] = data['dmaxs' + set[2]]
            elem['avgs'] = data['davgs' + set[2]]
        if set[2] == '':
            flag = flag + 1
            elem['children'] = []
            res.append(elem)

        else:
            res[flag]['children'].append(elem)
    return jsonify(json.dumps(res, cls=DecimalEncoder))


@bp.route('/checkout', methods=["GET"])
def checkout():
    """
    本函数用于数据库和EXCEL表相互校验
    """
    date = getdate()
    cdf_db = CalDailyForm.query.filter_by(date=date).first()
    cdf_excel = pd.read_excel(EXCEL_PATH, sheet_name='日报计算表', usecols=range(76), skiprows=range(3), header=None)
    x = (date - datetime.datetime(datetime.datetime.now().year, 1, 1)).days + 1
    condition = [cdf_db.dgp1 == cdf_excel.loc[x].values[21], cdf_db.donp1 == cdf_excel.loc[x].values[22],
                 cdf_db.doffp1 == cdf_excel.loc[x].values[23], cdf_db.dcp1 == cdf_excel.loc[x].values[24],
                 Decimal(cdf_db.dcl1) == realRound(cdf_excel.loc[x].values[25], 4),
                 cdf_db.dgp2 == cdf_excel.loc[x].values[26],
                 cdf_db.donp2 == cdf_excel.loc[x].values[27], cdf_db.doffp2 == cdf_excel.loc[x].values[28],
                 cdf_db.dcp2 == cdf_excel.loc[x].values[29],
                 Decimal(cdf_db.dcl2) == realRound(cdf_excel.loc[x].values[30], 4),
                 cdf_db.dgp == cdf_excel.loc[x].values[31], cdf_db.donp == cdf_excel.loc[x].values[32],
                 cdf_db.doffp == cdf_excel.loc[x].values[33], cdf_db.dcp == cdf_excel.loc[x].values[34],
                 Decimal(cdf_db.dcl) == realRound(cdf_excel.loc[x].values[35], 4),
                 Decimal(cdf_db.doffp31b) == realRound(cdf_excel.loc[x].values[36], 0),
                 cdf_db.doffp21b == cdf_excel.loc[x].values[37],
                 cdf_db.agp1 == cdf_excel.loc[x].values[38], cdf_db.aonp1 == cdf_excel.loc[x].values[39],
                 cdf_db.aoffp1 == cdf_excel.loc[x].values[40], cdf_db.acp1 == cdf_excel.loc[x].values[41],
                 Decimal(cdf_db.acl1) == realRound(cdf_excel.loc[x].values[42], 4),
                 cdf_db.agp2 == cdf_excel.loc[x].values[43],
                 cdf_db.aonp2 == cdf_excel.loc[x].values[44], cdf_db.aoffp2 == cdf_excel.loc[x].values[45],
                 cdf_db.acp2 == cdf_excel.loc[x].values[46],
                 Decimal(cdf_db.acl2) == realRound(cdf_excel.loc[x].values[47], 4),
                 cdf_db.agp == cdf_excel.loc[x].values[48], cdf_db.aonp == cdf_excel.loc[x].values[49],
                 cdf_db.aoffp == cdf_excel.loc[x].values[50], cdf_db.acp == cdf_excel.loc[x].values[51],
                 Decimal(cdf_db.acl) == realRound(cdf_excel.loc[x].values[52], 4),
                 cdf_db.mgp1 == cdf_excel.loc[x].values[53],
                 cdf_db.monp1 == cdf_excel.loc[x].values[54], cdf_db.moffp1 == cdf_excel.loc[x].values[55],
                 cdf_db.mcp1 == cdf_excel.loc[x].values[56],
                 Decimal(cdf_db.mcl1) == realRound(cdf_excel.loc[x].values[57], 4),
                 cdf_db.mgp2 == cdf_excel.loc[x].values[58], cdf_db.monp2 == cdf_excel.loc[x].values[59],
                 cdf_db.moffp2 == cdf_excel.loc[x].values[60], cdf_db.mcp2 == cdf_excel.loc[x].values[61],
                 Decimal(cdf_db.mcl2) == realRound(cdf_excel.loc[x].values[62], 4),
                 cdf_db.mgp == cdf_excel.loc[x].values[63],
                 cdf_db.monp == cdf_excel.loc[x].values[64], cdf_db.moffp == cdf_excel.loc[x].values[65],
                 cdf_db.mcp == cdf_excel.loc[x].values[66],
                 Decimal(cdf_db.mcl) == realRound(cdf_excel.loc[x].values[67], 4),
                 Decimal(cdf_db.offja311) == realRound(cdf_excel.loc[x].values[69], 0),
                 Decimal(cdf_db.offjr311) == realRound(cdf_excel.loc[x].values[71], 0),
                 Decimal(cdf_db.offja321) == realRound(cdf_excel.loc[x].values[73], 0),
                 Decimal(cdf_db.offjr321) == realRound(cdf_excel.loc[x].values[75], 0)]
    res = True
    for item in condition:
        res = True and item
    return jsonify(res)


@bp.route('/tooms', methods=["GET"])
def to_oms():
    """
    将填写数据写入oms报表
    """
    date = getdate()
    cdf = CalDailyForm.query.filter_by(date=date).first()
    workbook = openpyxl.load_workbook(OMS_PATH)
    worksheet = workbook['OMS日报']
    this_row_num = 0
    for row_num in range(1, worksheet.max_row):
        if worksheet.cell(row_num, 1).value == getdate() or worksheet.cell(row_num, 1).value in [None, '']:
            this_row_num = row_num
            break
    stop_time = get_stop_time(date)['sum']
    res = {
        'stop_time': float(realRound(stop_time, 2)),
        'installed_cap': 100,
        "fix_cap": float(realRound(stop_time / 24 * 2.5, 2)),
        "boot_cap": float(realRound(100 - stop_time / 24 * 2.5, 2)),
        "g_p": float(realRound(cdf.dgp / 10000, 2)),
        "on_p": float(realRound(cdf.donp / 10000, 2)),
        "used_p": float(realRound(cdf.dgp / 10000, 2) - realRound(cdf.donp / 10000, 2)),
        "blocked_p": float(realRound(get_lost_power(date)['sum'], 2)),
        "limited_p": float(realRound(cdf.dlp, 2)),
        'max_l': cdf.dmaxl,
        'min_l': cdf.dminl
    }
    worksheet.cell(this_row_num, 2, res['stop_time'])  # 维护+故障停机时间
    worksheet.cell(this_row_num, 3, res['installed_cap'])  # 装机容量
    worksheet.cell(this_row_num, 8, res['fix_cap'])  # 检修容量
    worksheet.cell(this_row_num, 4, res['boot_cap'])  # 开机容量
    worksheet.cell(this_row_num, 5, res['g_p'])  # 发电量
    worksheet.cell(this_row_num, 6, res['on_p'])  # 上网电量
    worksheet.cell(this_row_num, 7, res['used_p'])  # 场用电量
    worksheet.cell(this_row_num, 9, res['blocked_p'])  # 站内受阻电量
    worksheet.cell(this_row_num, 10, res['limited_p'])  # 站内受阻电量
    workbook.save(OMS_PATH)
    return jsonify([
        res
    ])


@bp.route('/toty', methods=["GET"])
def to_ty():
    """
    将填写数据写入桃园报表
    """
    date = getdate()
    cdf = CalDailyForm.query.filter_by(date=date).first()
    workbook = openpyxl.load_workbook(TY_PATH)
    ws_fdl = workbook['发电量统计']
    ws_fs = workbook['风速统计']
    ws_fjzt = workbook['风机状态统计']
    ws_xd = workbook['限电量统计']
    num = get_num(date)
    gz_num1 = 0
    gz_num2 = 0
    wh_num1 = 0
    wh_num2 = 0
    for x in range(0, 20):
        if num['wh'][x] != 0:
            wh_num1 = wh_num1 + 1
        elif num['gz'][x] != 0:
            gz_num1 = gz_num1 + 1
    for x in range(20, 40):
        if num['wh'][x] != 0:
            wh_num2 = wh_num2 + 1
        elif num['gz'][x] != 0:
            gz_num2 = gz_num2 + 1
    for row_num in range(1, ws_fdl.max_row):
        if ws_fdl.cell(row_num, 1).value == getdate():
            ws_fdl.cell(row_num, 2, realRound(cdf.dgp1 / 10000, 4))
            ws_fdl.cell(row_num, 3, realRound(cdf.dgp2 / 10000, 4))
            ws_fs.cell(row_num - 1, 2, realRound(cdf.davgs1, 2))
            ws_fs.cell(row_num - 1, 3, realRound(cdf.davgs2, 2))
            ws_fjzt.cell(row_num, 2, 20 - gz_num1 - wh_num1)
            ws_fjzt.cell(row_num, 3, gz_num1)
            ws_fjzt.cell(row_num, 4, wh_num1)
            ws_fjzt.cell(row_num, 5, 0)
            ws_fjzt.cell(row_num, 6, 20 - gz_num2 - wh_num2)
            ws_fjzt.cell(row_num, 7, gz_num2)
            ws_fjzt.cell(row_num, 8, wh_num2)
            ws_fjzt.cell(row_num, 9, 0)
            break
    pcs = PowerCut.query.filter(
        or_(PowerCut.start_time >= date, PowerCut.stop_time >= date)).all()
    changed_row_num = []
    today_lost_power = 0
    sum_lost_power = 0
    if pcs:
        for row_num in range(2, ws_xd.max_row):
            if ws_xd.cell(row_num, 3).value in [None, '']:
                for pc in pcs:
                    ws_xd.cell(row_num, 3, pc.start_time.strftime('%H:%M') + '-' + pc.stop_time.strftime('%H:%M'))
                    ws_xd.cell(row_num, 4, pc.start_time.strftime('%H:%M') + '-' + pc.stop_time.strftime('%H:%M'))
                    ws_xd.cell(row_num, 6, pc.lost_power1 + pc.lost_power2)
                    ws_xd.cell(row_num, 5, realRound((pc.stop_time - pc.start_time).seconds / 3600, 2))
                    changed_row_num.append(row_num)
                    today_lost_power = today_lost_power + ws_xd.cell(row_num, 6).value
                    row_num = row_num + 1
                ws_xd.cell(changed_row_num[0], 2).value = date.strftime('%Y.%m.%d')
                ws_xd.merge_cells(start_row=changed_row_num[0], start_column=2,
                                  end_row=changed_row_num[changed_row_num.__len__() - 1], end_column=2)
                ws_xd.cell(changed_row_num[0], 7).value = today_lost_power
                ws_xd.merge_cells(start_row=changed_row_num[0], start_column=7,
                                  end_row=changed_row_num[changed_row_num.__len__() - 1], end_column=7)
                row_num_pre = 0
                for merged_cell in ws_xd.merged_cells:
                    if merged_cell.min_row <= changed_row_num[0] - 1 <= merged_cell.max_row \
                            and 8 >= merged_cell.min_col and 8 <= merged_cell.max_col:
                        row_num_pre = merged_cell.min_row
                ws_xd.cell(changed_row_num[0], 8).value = sum_lost_power + today_lost_power
                ws_xd.merge_cells(start_row=changed_row_num[0], start_column=8,
                                  end_row=changed_row_num[changed_row_num.__len__() - 1], end_column=8)
                break
            else:
                sum_lost_power = sum_lost_power + ws_xd.cell(row_num, 6).value
    workbook.save(TY_PATH)
    response = jsonify()
    response.status_code = 200
    return response


@bp.route('/getbmz', methods=["GET"])
def get_bmz():
    """
    获取标码值
    """
    date = getdate()
    cdf = CalDailyForm.query.filter_by(date=date).first()
    if cdf == None:
        response = jsonify({})
        response.status_code = 202
    else:
        response = jsonify(cdf.to_dict())
    return response


@bp.route('/getdfs', methods=['GET'])
def get_dfs():
    """
    获取昨日及今日日报表数据
    """
    date = datetime.datetime.combine(datetime.date.today(), datetime.time(0, 0, 0)) + datetime.timedelta(-1)
    cdfs = CalDailyForm.query.filter(CalDailyForm.date >= date + datetime.timedelta(-1)) \
        .order_by(CalDailyForm.date.desc()) \
        .limit(2) \
        .all()
    if len(cdfs) == 0:
        import_cdf()
        cdfs = CalDailyForm.query.filter(CalDailyForm.date >= date + datetime.timedelta(-1)) \
            .order_by(CalDailyForm.date.desc()) \
            .limit(2) \
            .all()
        if len(cdfs) == 0:
            response = jsonify()
            response.status_code = 202
            return response
    data = []
    for cdf in cdfs:
        data.append({
            'date': cdf.date.strftime('%y/%m/%d'),
            'dgp': realRound(cdf.dgp / 10000, 4),
            'dlp': 0 if cdf.dlp in [None, ''] else cdf.dlp,
            'dlpl': 0 if cdf.dlp in [None, ''] else format(realRound(cdf.dlp / cdf.dgp * 100, 4), '0.2f') + '%',
            'donp': realRound(cdf.donp / 10000, 4),
            'doffp': realRound(cdf.doffp / 10000, 4),
            'dcp': realRound(cdf.dcp / 10000, 4),
            'dcl': format(realRound(cdf.dcp / cdf.dgp * 100, 4), '0.2f') + '%',
            'dmaxs': '' if cdf.dmaxs in [None, ''] else cdf.dmaxs,
            'dmins': '' if cdf.dmins in [None, ''] else cdf.dmins,
            'davgs': '' if cdf.davgs in [None, ''] else cdf.davgs,
        })
    return jsonify(json.dumps(data, cls=DecimalEncoder))


@bp.route('/getoms', methods=["GET"])
def get_oms():
    """
    将获取oms报表
    """
    date = getdate()
    cdf = CalDailyForm.query.filter_by(date=date).first()
    stop_time = get_stop_time(date)['sum']
    res = {
        'stop_time': float(realRound(stop_time, 2)),
        'installed_cap': 100,
        "fix_cap": float(realRound(stop_time / 24 * 2.5, 2)),
        "boot_cap": float(realRound(100 - stop_time / 24 * 2.5, 2)),
        "g_p": float(realRound(cdf.dgp / 10000, 2)),
        "on_p": float(realRound(cdf.donp / 10000, 2)),
        "used_p": float(realRound(cdf.dgp / 10000, 2) - realRound(cdf.donp / 10000, 2)),
        "blocked_p": float(realRound(get_lost_power(date)['sum'], 2)),
        "limited_p": float(realRound(cdf.dlp, 2)),
        'max_l': cdf.dmaxl,
        'min_l': cdf.dminl
    }
    return jsonify([
        res
    ])
